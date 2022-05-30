from cgi import test
from cgitb import text
from distutils.command.build_scripts import first_line_re
import os
from datetime import datetime
import sys
from time import time_ns
from unicodedata import name
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QWidget, QPushButton, QFileDialog
from PyQt5.QtGui import QIcon
from openpyxl import Workbook
from openpyxl import load_workbook

class App(QMainWindow):
    
    # !     Important
    # TODO  ToDo
    # *     Note
    # ?     Logic
    
    
    def __init__(self):
        super().__init__()
        self.title = 'Monthly Excel Report'
        self.left : int = 100
        self.top = 100
        self.width = 300
        self.height = 150
        self.setStyleSheet("""
                           background-color:#363636
                           """)
        
        self.initUI()
        
    def createTextblock(self,obj, int_pos_x, int_pos_y, int_size_x, int_size_y, style):
        obj.move(int_pos_x,int_pos_y)
        obj.resize(int_size_x,int_size_y)
        obj.setStyleSheet(style)
        
    def createQLabel(self,obj,name,int_x,int_y, style):
        obj.move(int_x,int_y)
        obj.setStyleSheet(style)
        obj.setText(name)
    

    def initUI(self):
        
        button_style = """
                color: #cfcfcf;
                background-color: #474747;
        """
        
        textbox_style ="""
                color: #cfcfcf;
                background-color: #474747;
        """
        
        qlabel_style = """
                color: #cfcfcf;
        """
        
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        self.statusBar().showMessage('Select master excel')
        self.setFixedSize(450,250)
        
        # Browse button
        button = QPushButton('browse',self)
        button.setToolTip('select excel file')
        button.move(20,180)
        button.clicked.connect(self.openFileNameDialog)
        button.setStyleSheet(button_style)
        
        # * ----------Textbox----------------------
        
        # Textboxes
        self.full_name = QLineEdit(self)
        self.user_account = QLineEdit(self)
        self.first_date = QLineEdit(self)
        self.last_date = QLineEdit(self)
        self.createTextblock(self.full_name,20,35,100,20,textbox_style)
        self.createTextblock(self.user_account,140,35,100,20,textbox_style)
        self.createTextblock(self.first_date,20,105,100,20,textbox_style)
        self.createTextblock(self.last_date,140,105,100,20,textbox_style)               
       
        # * --------------QLabels----------------
        
        # Qlabel name 
        person_name = QLabel(self)
        user_account = QLabel(self)
        first_date = QLabel(self)
        last_date = QLabel(self)
        info_zone= QLabel(self)
        info_zone.setFixedSize(200,140)
        
        info_zone_txt= """
        How to :
        
        1. Fill in the fields with the
        required data 
        Date format : 24.05.2022
        
        2. Then select the master excel 
        and relax
        """""
        self.createQLabel(info_zone,info_zone_txt,250,4,qlabel_style)
        self.createQLabel(person_name,'Full Name',20,4,qlabel_style)
        self.createQLabel(user_account,'Account',140,4,qlabel_style)
        self.createQLabel(first_date,'First Date',20,75,qlabel_style)
        self.createQLabel(last_date,'Last Date',140,75,qlabel_style)
                
        # *   Test Button
        # test_button = QPushButton('Test Button', self)
        # test_button.move(140,180)
        # test_button.clicked.connect(self.testButton)
        # test_button.setStyleSheet(button_style)
        
        self.show()
    
    
    # * Test Button
    # def testButton(self):
    #     first_date_obj = datetime.strptime(self.first_date.text(),'%d/%m/%Y')
    #     print(first_date_obj)
    #     print(type(first_date_obj))

        
    # * Open File Dialog Window Event
    def openFileNameDialog(self):     
        
        full_name = self.full_name.text()
        first_date_obj = datetime.strptime(self.first_date.text(),'%d.%m.%Y')
        last_date_obj = datetime.strptime(self.last_date.text(),'%d.%m.%Y')
        user_account = self.user_account.text()
        sheet_name = str(self.first_date.text())+"-"+str(self.last_date.text())
        
        excel_header = [
            "DATE",#0 #A
            "EC DOCS",#1 #B
            "IC DOCS",#2 #C
            "GC DOCS",#3 #D
            "EC TIME",#4 #E
            "IC TIME",#5 #F
            "GC TIME",#6 #G
            "EXTRA TIME",#7 #H
            "EC SPEED",#8 #I
            "IC SPEED",#9 #J
            "GC SPEED",#10 #K
            "TOTAL TIME",#11 #L
            "WORKED",#12 #M
            "MINUTES",#13 #N
            "MONTHLY EC SPEED",#14 #O
            "MONTHLY IC SPEED",#15 #P
            "MONTHLY GC SPEED",#16 #Q
        ]
        
        desktop = os.path.expanduser("~\Desktop\\") #path for current user desktop
        
        fileName, _ = QFileDialog.getOpenFileName(self,"Select Excel", "","Excel Files(*.xlsx)") #Grab File
        
        destination_workbook = Workbook() #inst destination_workbook
        destination_workbook.save(desktop+full_name+'.xlsx') #create destination_workbook
        destination_workbook.create_sheet(sheet_name,0)
        destination_worksheet = destination_workbook[sheet_name]
        
        destination_worksheet.append(excel_header)
        
        master_workbook = Workbook()
        master_workbook = load_workbook(fileName) #load_workbook from fileName
        master_worksheet = master_workbook.active #grabs active worksheet from master_workbook
        i=2
        for iteration in master_worksheet.iter_rows(min_row=2, values_only=True): #returns a tuple
            values_list = [0]*14
            excel_date = datetime.strptime(str(iteration[0]) ,'%d/%m/%Y')
            if ((first_date_obj <= excel_date <= last_date_obj) & (iteration[3] == user_account)):
                values_list[0] = iteration[0] # write date
                docs = iteration[4]
                time = iteration[5]
                extra_time = iteration[7]
                values_list[8] = "=IFERROR(B"+str(i)+"/E"+str(i)+",0)" #EC SPEED
                values_list[9] = "=IFERROR(C"+str(i)+"/F"+str(i)+",0)" #IC SPEED
                values_list[10] = "=IFERROR(D"+str(i)+"/G"+str(i)+",0)" #GC SPEED
                values_list[11] = "=SUM(E"+str(i)+"+""F"+str(i)+"+""G"+str(i)+"+""H"+str(i)+")" #TOTAL TIME
                values_list[12] = 8 #WORKED
                values_list[13] = "=(L"+str(i)+"-M"+str(i)+")" #MINUTES
                if(iteration[1]== "EC"):
                    values_list[1] = docs
                    values_list[4] = time
                    values_list[7] = extra_time
                    destination_worksheet.append(values_list)
                    i+=1
                if(iteration[1]== "IC"):
                    values_list[2] = docs
                    values_list[5] = time
                    values_list[7] = extra_time
                    destination_worksheet.append(values_list)
                    i+=1
                if(iteration[1]== "GC"):
                    values_list[3] = docs
                    values_list[6] = time
                    values_list[7] = extra_time
                    destination_worksheet.append(values_list)
                    i+=1
        destination_worksheet.append({
            "N":"=SUM(N2:N"+str(i-1)+")"
            })
        destination_worksheet.append({
            "O":"=SUM(B:B)/SUM(E:E)",
            "P":"=SUM(C:C)/SUM(F:F)",
            "Q":"=SUM(D:D)/SUM(G:G)"
            }) #MONTHLY EC SPEED MONTHLY IC SPEED MONTHLY GC SPEED

        destination_workbook.save(desktop+full_name+'.xlsx')
        self.statusBar().showMessage("File Saved")
        
        master_workbook.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('icon.png'))
    app.setStyle('Fusion')
    ex = App()
    sys.exit(app.exec_())
